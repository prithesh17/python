import openpyxl

lang = ['Kannnada', 'Telugu', 'Tamil', 'Malayalam', 'Konkani', 'Marathi']
capital = ['Bengaluru', 'Hyderabad', 'Chennai', 'Thiruvanthapuram', 'Panaji', 'Mumbai']
code = ['KA', 'TS', 'TN', 'KL', 'GA', 'MH']
state = ['Karnataka', 'Telangana', 'Tamilnadu', 'Kerala', 'Goa', 'Maharashtra']

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Language"

sheet.cell(1, 1, "State")
sheet.cell(1, 2, "Language")
sheet.cell(1, 3, "Capital")
sheet.cell(1, 4, "Code")

for i in range(2, 8):
    sheet.cell(i, 1, state[i - 2])
    sheet.cell(i, 2, lang[i - 2])
    sheet.cell(i, 3, capital[i - 2])
    sheet.cell(i, 4, code[i - 2])

wb.save("demo.xlsx")
print("Excel sheet created successfully")

wb = openpyxl.load_workbook("demo.xlsx")
sheet = wb.active

print("Reading from XL sheet:")
for row in sheet.iter_rows(1, sheet.max_row):
    print("\n")
    for cell in row:
        print(cell.value,end=' ')
